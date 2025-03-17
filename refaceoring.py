from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle
import os
from openpyxl.utils import column_index_from_string
import re
from openpyxl.styles import Font


def cut_and_paste_excel(file_path, src_range, dst_cell):
    """
    Вырезает диапазон ячеек и вставляет в указанную ячейку.

    :param file_path: Путь к Excel-файлу
    :param src_range: Строка с диапазоном (например, "B11:K18")
    :param dst_cell: Ячейка, куда вставить данные (например, "M5")
    """
    wb = load_workbook(file_path)
    ws = wb.active  # Активный лист

    # Получаем координаты
    src_start_col, src_start_row, src_end_col, src_end_row = get_range_coords(src_range)
    dst_start_col, dst_start_row = get_cell_coords(dst_cell)

    rows = src_end_row - src_start_row + 1
    cols = src_end_col - src_start_col + 1

    # Копируем данные в новую область
    for r in range(rows):
        for c in range(cols):
            src_cell = ws.cell(row=src_start_row + r, column=src_start_col + c)
            dst_cell = ws.cell(row=dst_start_row + r, column=dst_start_col + c)

            dst_cell.value = src_cell.value  # Копируем данные

            # Очищаем исходные ячейки
    for r in range(rows):
        for c in range(cols):
            ws.cell(row=src_start_row + r, column=src_start_col + c).value = None

    wb.save(file_path)


def get_range_coords(range_str):
    """ Преобразует строку диапазона (например, 'B11:K18') в числовые координаты. """
    start, end = range_str.split(":")
    start_col, start_row = split_cell(start)
    end_col, end_row = split_cell(end)
    return start_col, start_row, end_col, end_row


def get_cell_coords(cell_str):
    """ Преобразует строку ячейки ('M5') в координаты (номер столбца, номер строки). """
    col, row = split_cell(cell_str)
    return col, row


def split_cell(cell_str):
    """ Разделяет строку ячейки на букву столбца и номер строки. """
    match = re.match(r"([A-Z]+)(\d+)", cell_str, re.I)
    if match:
        col, row = match.groups()
        return column_index_from_string(col), int(row)
    raise ValueError(f"Некорректный формат ячейки: {cell_str}")


def clear_row(file_path, sheet_name, row_num):
    """
    Очищает все ячейки в указанной строке.

    :param file_path: Путь к Excel-файлу
    :param sheet_name: Название листа
    :param row_num: Номер строки (например, 5)
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]  # Выбираем нужный лист

    # Очищаем значения всех ячеек в строке
    for cell in ws[row_num]:
        cell.value = None

    wb.save(file_path)


def clear_cells_in_range(file_path, sheet_name, cell_range):
    # Загружаем рабочую книгу
    workbook = load_workbook(file_path)

    # Получаем лист по имени
    sheet = workbook[sheet_name]

    # Очищаем ячейки в указанном диапазоне
    for row in sheet[cell_range]:
        for cell in row:
            cell.value = None  # Очищаем значение ячейки

    # Сохраняем изменения
    workbook.save(file_path)


# Пример использования


def delete_row(file_path, sheet_name, row_number):
    # Загружаем рабочую книгу
    workbook = load_workbook(file_path)

    # Получаем лист по имени
    sheet = workbook[sheet_name]

    # Удаляем строку (второй аргумент указывает на количество строк для удаления, по умолчанию 1)
    sheet.delete_rows(row_number)

    # Сохраняем изменения
    workbook.save(file_path)


def merge_cells(file_path, cell_range, value=None):
    """
    Функция для объединения ячеек в Excel.

    :param file_path: Лист (worksheet) в openpyxl
    :param cell_range: Диапазон ячеек в формате 'A1:C3'
    :param value: Значение, которое нужно записать в первую ячейку диапазона (по умолчанию None)
    """
    wb = load_workbook(file_path)
    ws = wb.active
    ws.merge_cells(cell_range)  # Объединяем ячейки
    if value:
        top_left_cell = cell_range.split(":")[0]  # Получаем первую ячейку диапазона
        ws[top_left_cell] = value  # Записываем текст
        ws[top_left_cell].alignment = Alignment(horizontal="center", vertical="center")  # Центрируем текст
    wb.save(file_path)


def clean_excel(file_path, output_file, target_col=2):
    """
    Функция для удаления нечётных столбцов (кроме первого) и разъединения объединённых ячеек.

    :param file_path: Путь к исходному файлу Excel
    :param output_file: Путь для сохранения изменённого файла
    :param target_col: Столбец, с которого начинается удаление (по умолчанию 2 - столбец 'B')
    """
    # Загружаем Excel-файл
    wb = load_workbook(file_path)
    ws = wb.active  # Выбираем активный лист

    # Разъединяем все объединённые ячейки
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    # Определяем максимальное количество столбцов
    max_col = ws.max_column

    # Удаляем нечётные столбцы, начиная со второго
    for col in range(max_col, target_col, -1):  # Идём с конца, чтобы не сдвигались индексы
        if col % 2 == 1:  # Проверяем, нечётный ли номер столбца
            ws.delete_cols(col)

    # Сохраняем изменения
    wb.save(output_file)


def get_cell_value(file_path, cell):
    # Загружаем рабочую книгу
    wb = load_workbook(file_path)

    # Получаем лист по имени
    ws = wb.active

    # Получаем значение ячейки
    value = ws[cell].value
    return value


def change_cell_format(file_path, cell_ranges):
    # Загружаем рабочую книгу и выбираем лист
    wb = load_workbook(file_path)
    ws = wb.active
    wb.guess_types = True
    numberStyle = NamedStyle(name='numberStyle', number_format='0.00')

    for cell_range in cell_ranges:
        # Проходим по указанным диапазонам ячеек
        for row in ws[cell_range]:
            for i in row:
                i.style = numberStyle
                i.font = Font(name='Times New Roman', size=9)

        # Сохраняем изменения в файле
    wb.save(file_path)


def is_refactoring(name_file):
    result = (get_cell_value(name_file, 'AD4'))
    return result


def font(file_path, cell_range, size=9):
    # Загружаем рабочую книгу и выбираем лист
    wb = load_workbook(file_path)
    ws = wb.active

    # Проходим по указанным диапазонам ячеек
    # Проходим по указанному диапазону ячеек
    for row in ws[cell_range]:
        for cell in row:
            cell.font = Font(name='Calibri', size=11)  # Устанавливаем шрифт для каждой ячейки
    wb.save(file_path)


def refactoring_file(name_file):
    if not is_refactoring(name_file):
        value = get_cell_value(name_file, 'B3')
        y_ = get_cell_value(name_file, 'B25')
        z_ = get_cell_value(name_file, 'B47')
        print(value, y_, z_)
        clean_excel(name_file, name_file)
        # Получить длину
        cut_and_paste_excel(name_file, "B11:K17", "L4")
        cut_and_paste_excel(name_file, "B18:K24", "V4")
        clear_row(name_file, "Sheet1", 11)  # Очистит строку 11
        cut_and_paste_excel(name_file, "A25:K32", "A11")
        cut_and_paste_excel(name_file, "B33:K39", "L12")
        cut_and_paste_excel(name_file, "B40:K46", "V12")
        clear_row(name_file, "Sheet1", 19)  # Очистит строку 11
        cut_and_paste_excel(name_file, "A47:K54", "A19")
        cut_and_paste_excel(name_file, "B55:K61", "L20")
        cut_and_paste_excel(name_file, "B62:K68", "V20")
        clear_cells_in_range(name_file, 'Sheet1', 'A27:AA100')
        delete_row(name_file, 'Sheet1', 2)  # Удаляет вторую строку на листе 'Sheet1'
        merge_cells(name_file, 'A1:AE1',
                    'Значения виброскоростей, мкм/с в 1/3 октавной полосе со среднегеометрической частотой, Гц')
        # if value == 'Ось Y':
        # merge_cells(name_file, 'A2:AD2', 'Ось Y')
        # merge_cells(name_file, 'A10:AD10', 'Ось Z')
        # else:
        merge_cells(name_file, 'A2:AE2', value)
        merge_cells(name_file, 'A10:AE10', y_)
        merge_cells(name_file, 'A18:AE18', z_)

        change_cell_format(name_file, ['A4:AE9', 'A12:AE17', 'A20:AE25'])
        font(name_file, 'A1:AE25')
    else:
        print(f'Файл: "{name_file}" уже изменен!')


def main():
    path_dir = r'C:\Users\Fad\PycharmProjects\EXEL_REFACTORING\Новая папка (2)'
    files = os.listdir(path_dir)
    count_file = len(files)
    count_step = 0
    for i in files:
        count_step += 1
        print(str(count_step) + ' из ' + str(count_file))
        refactoring_file(path_dir + '/' + i)

    print('\nГотово')


main()
